
import streamlit as st
import pandas as pd
import numpy as np
import io

from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus import PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.markdown(
    """
    <style>
        .block-container {
            padding-top: 2.5rem;
        }
    </style>
    """,
    unsafe_allow_html=True
)

    
# =========================
# FORMAT NUMERIC FUNCTION
# =========================
def format_number(value):
    if pd.isna(value):
        return ""
    if isinstance(value, (int, np.integer)):
        return f"{value:,}"
    if isinstance(value, (float, np.floating)):
        return f"{value:,.2f}"
    return str(value)


# =========================
# AGGRID FUNCTION
# =========================
def show_aggrid(df, key_name):

    gb = GridOptionsBuilder.from_dataframe(df)

    # Pagination
    gb.configure_pagination(
        paginationAutoPageSize=False,
        paginationPageSize=25
    )

    # Default column behavior
    gb.configure_default_column(
        filter=True,
        sortable=True,
        resizable=True,
        floatingFilter=True,
        wrapText=False,
        autoHeight=False,
        minWidth=120,
    )

    # Sidebar (filter & column selector)
    gb.configure_side_bar()
    
    numeric_cols = df.select_dtypes(include=["int64", "float64"]).columns

    for col in numeric_cols:
        gb.configure_column(
            col,
            type=["numericColumn"],
            valueFormatter="x.toLocaleString()"
        )

    # Freeze kolom Kecamatan jika ada
    if "Kecamatan" in df.columns:
        gb.configure_column(
            "Kecamatan",
            pinned="left"
        )
    
    # Freeze kolom pertama (index 0)
    if len(df.columns) > 0:
        first_col = df.columns[0]

        gb.configure_column(
            first_col,
            pinned="left"
        )

    # Grid options
    gb.configure_grid_options(
        domLayout="normal",
        enableRangeSelection=True
    )

    grid_options = gb.build()

    AgGrid(
        df,
        gridOptions=grid_options,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=True,
        enable_enterprise_modules=False,
        theme="streamlit",
        key=key_name,
        height=500
    )



# =========================
# STREAMLIT UI
# =========================
st.title("Pivot Progres SYNC DAPODIK")

uploaded_file = st.file_uploader("Upload File Excel", type=["xlsx"])

if uploaded_file:

    # =========================
    # BACA FILE
    # =========================
    dapodik = pd.read_excel(uploaded_file, sheet_name="Master")

    dapodik_filtered = dapodik.loc[~dapodik["BP"].isin(["SMA", "SMK", "SLB"])].copy()
    dapodik_filtered["GTK"] = (
            dapodik_filtered["Guru"] + dapodik_filtered["Tendik"]
        )
    dapodik_filtered["Status_SYNC"] = dapodik_filtered["Last Sync"].apply(
            lambda x: "Belum" if x == "Belum Kirim" else "Sudah")


    dapodik_filtered["BP"] = pd.Categorical(
        dapodik_filtered["BP"],
        categories=["SPS", "PKBM", "TPA", "KB", "TK", "SD", "SMP", "SKB"],
        ordered=True
    )

    dapodik_filtered["Status"] = pd.Categorical(
        dapodik_filtered["Status"],
        categories=["Negeri", "Swasta"],
        ordered=True
    )
    
    dapodik_filtered["Sudah SYNC"] = (dapodik_filtered["Status_SYNC"] == "Sudah").astype(int)
    dapodik_filtered["Belum SYNC"] = (dapodik_filtered["Status_SYNC"] == "Belum").astype(int)
    
    dapodik_filtered = dapodik_filtered.rename(columns={"NPSN": "SP"}, level=0)
    
    
    # =========================
    # FILTER JENJANG SMP
    # =========================
    dapodik_filtered_SMP = dapodik.loc[dapodik["BP"].isin(["SMP"])].copy()
    
    dapodik_filtered_SMP["Status_SYNC"] = dapodik_filtered_SMP["Last Sync"].apply(
        lambda x: "Belum" if x == "Belum Kirim" else "Sudah"
    )
    dapodik_filtered_SMP["GTK"] = (
        dapodik_filtered_SMP["Guru"] + dapodik_filtered_SMP["Tendik"]
    )
    
    dapodik_filtered_SMP["Status"] = pd.Categorical(
        dapodik_filtered_SMP["Status"],
        categories=["Negeri", "Swasta"],
        ordered=True
    )
    
    dapodik_filtered_SMP["Sudah SYNC"] = (dapodik_filtered_SMP["Status_SYNC"] == "Sudah").astype(int)
    dapodik_filtered_SMP["Belum SYNC"] = (dapodik_filtered_SMP["Status_SYNC"] == "Belum").astype(int)
    
    dapodik_filtered_SMP = dapodik_filtered_SMP.rename(columns={"NPSN": "SP"}, level=0)
    
    # =========================
    # PIVOT
    # =========================
    per_kec = dapodik_filtered.pivot_table(
        values=["SP", "Sudah SYNC", "Belum SYNC", "PD", "Rombel", "Guru", "Tendik", "GTK"], 
        index="Kecamatan", 
        columns="Status", 
        aggfunc={
            "SP" : "nunique", 
            "Sudah SYNC" : "count",
            "Belum SYNC" : "count",
            "PD" : "sum", 
            "Rombel" : "sum",
            "Guru" : "sum",
            "Tendik" : "sum",
            "GTK" : "sum"
        }, 
        fill_value=0, 
        margins=True, 
        margins_name="Total", 
        observed=False
    )
    
    per_kec_NS = dapodik_filtered.pivot_table(
        values=["SP", "PD", "Rombel", "Guru", "Tendik", "GTK"], 
        index=["Kecamatan"], 
        columns=["Status"], 
        aggfunc={
            "SP" : "nunique", 
            "PD" : "sum", 
            "Rombel" : "sum",
            "Guru" : "sum",
            "Tendik" : "sum",
            "GTK" : "sum"
        }, 
        fill_value=0, 
        margins=True,
        margins_name="Total", 
        observed=False
        )

    per_bp_NS = dapodik_filtered.pivot_table(
        values=["SP", "PD", "Rombel", "Guru", "Tendik", "GTK"], 
        index=["BP"], 
        columns=["Status"], 
        aggfunc={
            "SP" : "nunique",
            "PD" : "sum", 
            "Rombel" : "sum",
            "Guru" : "sum",
            "Tendik" : "sum",
            "GTK" : "sum"
        }, 
        fill_value=0, 
        margins=True,
        margins_name="Total", 
        observed=False
        )
    
    # =========================
    # PIVOT JENJANG SMP
    # =========================
    
    per_kec_SMP = dapodik_filtered_SMP.pivot_table(
        values=["SP", "Sudah SYNC", "Belum SYNC", "PD", "Rombel", "Guru", "Tendik", "GTK"], 
        index=["Kecamatan"],
        columns=["Status"], 
        aggfunc={
            "SP" : "nunique", 
            "Sudah SYNC" : "count", 
            "Belum SYNC" : "count", 
            "PD" : "sum", 
            "Rombel" : "sum",
            "Guru" : "sum",
            "Tendik" : "sum",
            "GTK" : "sum"
        }, 
        fill_value=0, 
        margins=True,
        margins_name="Total", 
        observed=False
        )

    per_kec_SP_SMP = dapodik_filtered_SMP.pivot_table(
        values=["SP", "Sudah SYNC", "Belum SYNC"], 
        index=["Kecamatan"],
        columns=["Status"], 
        aggfunc={
            "SP" : "nunique", 
            "Sudah SYNC" : "count", 
            "Belum SYNC" : "count"
        }, 
        fill_value=0, 
        margins=True,
        margins_name="Total", 
        observed=False
        )
    
    per_kec_PDRombel_SMP = dapodik_filtered_SMP.pivot_table(
        values=["PD", "Rombel"], 
        index=["Kecamatan"],
        columns=["Status"], 
        aggfunc={
            "PD" : "sum", 
            "Rombel" : "sum"
        }, 
        fill_value=0, 
        margins=True,
        margins_name="Total", 
        observed=False
        )
    
    per_kec_GTK_SMP = dapodik_filtered_SMP.pivot_table(
        values=["Guru", "Tendik", "GTK"], 
        index=["Kecamatan"],
        columns=["Status"], 
        aggfunc={
            "Guru" : "sum",
            "Tendik" : "sum",
            "GTK" : "sum"
        }, 
        fill_value=0, 
        margins=True,
        margins_name="Total", 
        observed=False
        )

    per_kec_ALL_SMP = dapodik_filtered_SMP.pivot_table(
        values=["SP", "PD", "Rombel", "Guru", "Tendik", "GTK"], 
        index=["Kecamatan"],
        columns=["Status"], 
        aggfunc={
            "SP" : "nunique",
            "PD" : "sum", 
            "Rombel" : "sum",
            "Guru" : "sum",
            "Tendik" : "sum",
            "GTK" : "sum"
        }, 
        fill_value=0, 
        margins=True,
        margins_name="Total", 
        observed=False
        )

    st.success("Data berhasil diproses!")
    
    # st.write("Kolom sebelum export:", per_kec_final.columns.tolist()) DEBUG

    
    # =========================
    # KPI DASHBOARD
    # =========================

    total_sp = dapodik_filtered["SP"].nunique()
    
    total_sudah = dapodik_filtered["Sudah SYNC"].sum()

    total_belum = total_sp - total_sudah

    persen_sync = (total_sudah / total_sp * 100) if total_sp > 0 else 0

    col1, col2, col3, col4 = st.columns(4)

    col1.metric("Total Satuan Pendidikan", f"{total_sp:,}")
    col2.metric("Sudah SYNC", f"{total_sudah:,}")
    col3.metric("Belum SYNC", f"{total_belum:,}")
    col4.metric("Persentase SYNC", f"{persen_sync:.2f}%")
    
    # =========================
    # GRAFIK PROGRES SYNC
    # =========================

    st.subheader("📊 Grafik Progres SYNC per Kecamatan")

    chart_data = (
        dapodik_filtered
        .groupby("Kecamatan")
        .agg({
            "SP": "nunique",
            "Sudah SYNC": "sum",
            "Belum SYNC": "sum"
        })
    )

    # Hindari double count jika ada duplikasi NPSN
    chart_data["Sudah SYNC"] = (
        dapodik_filtered
        .loc[dapodik_filtered["Sudah SYNC"] == 1]
        .groupby("Kecamatan")["SP"]
        .nunique()
    )

    chart_data["Belum SYNC"] = (
        dapodik_filtered
        .loc[dapodik_filtered["Belum SYNC"] == 1]
        .groupby("Kecamatan")["SP"]
        .nunique()
    )

    chart_data = chart_data.fillna(0).astype(int)

    st.bar_chart(chart_data[["Sudah SYNC", "Belum SYNC"]])
    
    # =========================
    # RANKING KECAMATAN
    # =========================

    ranking = (
        dapodik_filtered
        .groupby("Kecamatan")
        .apply(lambda df: pd.Series({
            "Total_SP": df["SP"].nunique(),
            "Total_Sudah": df.loc[df["Belum SYNC"] == 0, "SP"].nunique()
        }))
        .reset_index()
    )

    ranking["Total_Belum"] = ranking["Total_SP"] - ranking["Total_Sudah"]

    ranking["Persentase_SYNC"] = (
        ranking["Total_Sudah"] / ranking["Total_SP"] * 100
    ).round(2)

    ranking = ranking.sort_values(
        "Persentase_SYNC", ascending=False
    )
    
    st.dataframe(
        ranking[[
            "Kecamatan",
            "Total_SP",
            "Total_Sudah",
            "Total_Belum",
            "Persentase_SYNC"
        ]],
        use_container_width=True
    )

    
    def prepare_table(df, index_name):
        df = df.copy()

        # Flatten jika MultiIndex
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [
                f"{col1}_{col2}" if col2 else col1
                for col1, col2 in df.columns
            ]

        # Reset index
        df = df.reset_index()

        # Pastikan numerik
        for col in df.columns:
            if col != index_name:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

        return df


    per_kec = (
        prepare_table(per_kec, "Kecamatan")
        [["Kecamatan", "SP_Total", "Sudah SYNC_Total", "Belum SYNC_Total", "PD_Total", "Rombel_Total", 
        "Guru_Total", "Tendik_Total", "GTK_Total"]]
        .rename(columns={
            "SP_Total": "SP",
            "Sudah SYNC_Total": "Sudah SYNC",
            "Belum SYNC_Total": "Belum SYNC",
            "PD_Total": "PD",
            "Rombel_Total": "Rombel",
            "Guru_Total": "Guru",
            "Tendik_Total": "Tendik",
            "GTK_Total": "GTK"
            })
    )
    
    per_kec_NS = (
        prepare_table(per_kec_NS, "Kecamatan")
        [["Kecamatan", 
        "SP_Negeri", "SP_Swasta", "SP_Total",
        "PD_Negeri", "PD_Swasta", "PD_Total",
        "Rombel_Negeri", "Rombel_Swasta", "Rombel_Total", 
        "Guru_Negeri", "Guru_Swasta", "Guru_Total", 
        "Tendik_Negeri", "Tendik_Swasta", "Tendik_Total", 
        "GTK_Negeri", "GTK_Swasta", "GTK_Total"]]
        .rename(columns={
            "SP_Negeri": "SP Negeri", "SP_Swasta": "SP Swasta", "SP_Total": "Jml SP",
            "PD_Negeri": "PD Negeri", "PD_Swasta": "PD Swasta", "PD_Total": "Jml PD",
            "Rombel_Negeri": "Rombel Negeri", "Rombel_Swasta": "Rombel Swasta", "Rombel_Total": "Jml Rombel",
            "Guru_Negeri": "Guru Negeri", "Guru_Swasta": "Guru Swasta", "Guru_Total": "Jml Guru",
            "Tendik_Negeri": "Tendik Negeri", "Tendik_Swasta": "Tendik Swasta", "Tendik_Total": "Jml Tendik",
            "GTK_Negeri": "GTK Negeri", "GTK_Swasta": "GTK Swasta", "GTK_Total": "Jml GTK"
            })
    )
    
    per_bp_NS = (
        prepare_table(per_bp_NS, "BP")
        [["BP", 
        "SP_Negeri", "SP_Swasta", "SP_Total",
        "PD_Negeri", "PD_Swasta", "PD_Total",
        "Rombel_Negeri", "Rombel_Swasta", "Rombel_Total", 
        "Guru_Negeri", "Guru_Swasta", "Guru_Total", 
        "Tendik_Negeri", "Tendik_Swasta", "Tendik_Total", 
        "GTK_Negeri", "GTK_Swasta", "GTK_Total"]]
        .rename(columns={
            "SP_Negeri": "SP Negeri", "SP_Swasta": "SP Swasta", "SP_Total": "Jml SP",
            "PD_Negeri": "PD Negeri", "PD_Swasta": "PD Swasta", "PD_Total": "Jml PD",
            "Rombel_Negeri": "Rombel Negeri", "Rombel_Swasta": "Rombel Swasta", "Rombel_Total": "Jml Rombel",
            "Guru_Negeri": "Guru Negeri", "Guru_Swasta": "Guru Swasta", "Guru_Total": "Jml Guru",
            "Tendik_Negeri": "Tendik Negeri", "Tendik_Swasta": "Tendik Swasta", "Tendik_Total": "Jml Tendik",
            "GTK_Negeri": "GTK Negeri", "GTK_Swasta": "GTK Swasta", "GTK_Total": "Jml GTK"
            })
    )
    
    
    per_kec_SMP = (
        prepare_table(per_kec_SMP, "Kecamatan")
        [["Kecamatan", "SP_Total", "Sudah SYNC_Total", "Belum SYNC_Total", "PD_Total", "Rombel_Total", 
        "Guru_Total", "Tendik_Total", "GTK_Total"]]
        .rename(columns={
            "SP_Total": "SP",
            "Sudah SYNC_Total": "Sudah SYNC",
            "Belum SYNC_Total": "Belum SYNC",
            "PD_Total": "PD",
            "Rombel_Total": "Rombel",
            "Guru_Total": "Guru",
            "Tendik_Total": "Tendik",
            "GTK_Total": "GTK"
            })
    )
    
    per_kec_SP_SMP = (
        prepare_table(per_kec_SP_SMP, "Kecamatan")
        [["Kecamatan", "SP_Total", "SP_Negeri", "SP_Swasta", "Sudah SYNC_Total", "Belum SYNC_Total"]]
        .rename(columns={
            "SP_Total": "Jml SP", "SP_Negeri": "SP Negeri", "SP_Swasta": "SP Swasta",
            "Sudah SYNC_Total": "Sudah SYNC", "Belum SYNC_Total": "Belum SYNC"
            })
    )
    
    per_kec_PDRombel_SMP = (
        prepare_table(per_kec_PDRombel_SMP, "Kecamatan")
        [["Kecamatan", "PD_Negeri", "PD_Swasta", "PD_Total", "Rombel_Negeri", "Rombel_Swasta", "Rombel_Total"]]
        .rename(columns={
            "PD_Negeri": "PD Negeri", "PD_Swasta": "PD Swasta", "PD_Total": "Jml PD",
            "Rombel_Negeri": "Rombel Negeri", "Rombel_Swasta": "Rombel Swasta", "Rombel_Total": "Jml Rombel",
            })
    )
    
    per_kec_GTK_SMP = (
        prepare_table(per_kec_GTK_SMP, "Kecamatan")
        [["Kecamatan", "GTK_Total", "Guru_Negeri", "Guru_Swasta", "Guru_Total", "Tendik_Negeri", "Tendik_Swasta", "Tendik_Total"]]
        .rename(columns={
            "GTK_Total": "Jml GTK",
            "Guru_Negeri": "Guru Negeri", "Guru_Swasta": "Guru Swasta", "Guru_Total": "Jml Guru",
            "Tendik_Negeri": "Tendik Negeri", "Tendik_Swasta": "Tendik Swasta", "Tendik_Total": "Jml Tendik",
            })
    )
    
    per_kec_ALL_SMP = (
        prepare_table(per_kec_ALL_SMP, "Kecamatan")
        [["Kecamatan", 
        "SP_Negeri", "SP_Swasta", 
        "PD_Negeri", "PD_Swasta", 
        "Rombel_Negeri", "Rombel_Swasta", 
        "Guru_Negeri", "Guru_Swasta", 
        "Tendik_Negeri", "Tendik_Swasta", 
        "SP_Total", "PD_Total", "Rombel_Total", "Guru_Total", "Tendik_Total", "GTK_Total", ]]
        .rename(columns={
            "SP_Negeri": "SP Negeri", "SP_Swasta": "SP Swasta",
            "PD_Negeri": "PD Negeri", "PD_Swasta": "PD Swasta",
            "Rombel_Negeri": "Rombel Negeri", "Rombel_Swasta": "Rombel Swasta",
            "Guru_Negeri": "Guru Negeri", "Guru_Swasta": "Guru Swasta", 
            "Tendik_Negeri": "Tendik Negeri", "Tendik_Swasta": "Tendik Swasta",
            "SP_Total": "Jml SP", "PD_Total": "Jml PD", "Rombel_Total": "Jml Rombel", 
            "Guru_Total": "Jml Guru", "Tendik_Total": "Jml Tendik", "GTK_Total": "Jml GTK",            
            })
    )


    tab1, tab2, tab3 = st.tabs([
        "📊 Rekap Per Kecamatan",
        "🏫 Rekap Negeri/Swasta Per Kecamatan",
        "📌 Rekap Negeri/Swasta Per BP"
    ])

    with tab1:
        show_aggrid(per_kec, "grid_kec")

    with tab2:
        show_aggrid(per_kec_NS, "grid_kec_ns")

    with tab3:
        show_aggrid(per_bp_NS, "grid_bp_ns")


    tab_smp_1, tab_smp_2, tab_smp_3, tab_smp_4, tab_smp_5 = st.tabs([
        "📊 DAPODIK SMP",
        "🏫 SP dan SYNC SMP",
        "📌 PD dan Rombel SMP",
        "📌 Progres GTK SMP",
        "📌 ALL DAPODIK SMP"
    ])

    with tab_smp_1:
        show_aggrid(per_kec_SMP, "grid_kec_smp")

    with tab_smp_2:
        show_aggrid(per_kec_SP_SMP, "grid_kec_sp_smp")

    with tab_smp_3:
        show_aggrid(per_kec_PDRombel_SMP, "grid_kec_pdrombel_smp")

    with tab_smp_4:
        show_aggrid(per_kec_GTK_SMP, "grid_kec_gtk_smp")

    with tab_smp_5:
        show_aggrid(per_kec_ALL_SMP, "grid_kec_all_smp")


    # =========================================
    # GLOBAL STYLE CONFIG
    # =========================================
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(size=14, bold=True)
    TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    NUMBER_FORMAT_INTEGER = '#,##0'
    NUMBER_FORMAT_DECIMAL = '#,##0.00'
    NUMBER_FORMAT_PERCENT = '0.00%'


    # =========================================
    # STYLE TABLE FUNCTION
    # =========================================
    def style_table(ws, start_row, df, title=None):
        """
        start_row : baris pertama tabel (1-based Excel)
        df        : dataframe
        """

        n_cols = len(df.columns)
        header_row = start_row
        data_start = header_row + 1
        data_end = data_start + len(df) - 1
        total_row = data_end

        # =========================
        # TITLE
        # =========================
        if title:
            ws.insert_rows(start_row)
            ws.cell(row=start_row, column=1, value=title)
            ws.cell(row=start_row, column=1).font = TITLE_FONT
            ws.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=start_row,
                end_column=n_cols
            )
            header_row += 1
            data_start += 1
            data_end += 1
            total_row += 1

        # =========================
        # HEADER STYLE
        # =========================
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # =========================
        # DATA + TOTAL STYLE
        # =========================
        for row in ws.iter_rows(
            min_row=data_start,
            max_row=data_end,
            min_col=1,
            max_col=n_cols
            
        ):
            for i, cell in enumerate(row):
                col_name = df.columns[i]
                
                cell.border = THIN_BORDER                
                
                if i == 0:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="right")
                
                # =========================
                # FORMAT NUMERIC OTOMATIS
                # =========================
                if pd.api.types.is_integer_dtype(df[col_name]):
                    cell.number_format = NUMBER_FORMAT_INTEGER

                elif pd.api.types.is_float_dtype(df[col_name]):

                    # Jika kolom persentase
                    if "%" in col_name.lower() or "persen" in col_name.lower():
                        cell.number_format = NUMBER_FORMAT_PERCENT
                    else:
                        cell.number_format = NUMBER_FORMAT_DECIMAL

        # Highlight TOTAL row
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)


        # =========================
        # AUTO WIDTH (Only Table Range)
        # =========================
        for col in range(1, n_cols + 1):
            column_letter = get_column_letter(col)
            max_len = 0

            for row in range(header_row, data_end + 1):
                value = ws.cell(row=row, column=col).value
                if value:
                    max_len = max(max_len, len(str(value)))

            ws.column_dimensions[column_letter].width = max_len + 3


    # =========================================
    # WRITE MULTIPLE TABLES
    # =========================================
    def write_tables(writer, sheet_name, tables, spacing=4):

        ws = writer.book.create_sheet(sheet_name)
        writer.sheets[sheet_name] = ws

        current_row = 1

        for df, title in tables:

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=current_row - 1
            )

            style_table(
                ws=ws,
                start_row=current_row,
                df=df,
                title=title
            )

            # Update posisi berikutnya
            current_row += len(df) + spacing + 2


    def flatten_columns(df):
        df = df.copy()
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = ["_".join(col).strip() for col in df.columns]
        return df


    def select_columns(df, cols):
        df = flatten_columns(df)
        return df.reindex(columns=cols)

    # ======================
    # MASTER EXPORT CONFIG
    # ======================

    master_exports = {
        "Rekapitulasi Progres SYNC DAPODIK per Kecamatan": (
            per_kec,
            ["Kecamatan","SP","Sudah SYNC","Belum SYNC","PD","Rombel","Guru","Tendik","GTK"]
        ),
        "Rekapitulasi SYNC Negeri/Swasta per Kecamatan": (
            per_kec_NS,
            ["Kecamatan","SP Negeri","SP Swasta","Jml SP",
            "PD Negeri","PD Swasta","Jml PD",
            "Rombel Negeri","Rombel Swasta","Jml Rombel",
            "Guru Negeri","Guru Swasta","Jml Guru",
            "Tendik Negeri","Tendik Swasta","Jml Tendik",
            "GTK Negeri","GTK Swasta","Jml GTK"]
        ),
        "Rekapitulasi SYNC Negeri/Swasta per Jenjang": (
            per_bp_NS,
            ["BP","SP Negeri","SP Swasta","Jml SP",
            "PD Negeri","PD Swasta","Jml PD",
            "Rombel Negeri","Rombel Swasta","Jml Rombel",
            "Guru Negeri","Guru Swasta","Jml Guru",
            "Tendik Negeri","Tendik Swasta","Jml Tendik",
            "GTK Negeri","GTK Swasta","Jml GTK"]
        )
    }

    smp_exports = {
        "Eksplorasi Progres SMP per Kecamatan": (
            per_kec_SMP,
            ["Kecamatan","SP","Sudah SYNC","Belum SYNC",
            "PD","Rombel","Guru","Tendik","GTK"]
        ),
        "Eksplorasi SP & SYNC SMP": (
            per_kec_SP_SMP,
            ["Kecamatan","Jml SP","SP Negeri","SP Swasta",
            "Sudah SYNC","Belum SYNC"]
        ),
        "Eksplorasi PD & Rombel SMP": (
            per_kec_PDRombel_SMP,
            ["Kecamatan","PD Negeri","PD Swasta","Jml PD",
            "Rombel Negeri","Rombel Swasta","Jml Rombel"]
        ),
        "Eksplorasi GTK SMP": (
            per_kec_GTK_SMP,
            ["Kecamatan","Jml GTK","Guru Negeri","Guru Swasta",
            "Jml Guru","Tendik Negeri","Tendik Swasta","Jml Tendik"]
        ),
        "Eksplorasi ALL SMP": (
            per_kec_ALL_SMP,
            ["Kecamatan","SP Negeri","SP Swasta","PD Negeri","PD Swasta",
            "Rombel Negeri","Rombel Swasta",
            "Guru Negeri","Guru Swasta","Jml Guru",
            "Tendik Negeri","Tendik Swasta",
            "Jml SP","Jml PD","Jml Rombel","Jml Tendik","Jml GTK"]
        )
    }

    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:

        write_tables(
            writer,
            "Rekap_Master",
            [(select_columns(df, cols), title)
            for title, (df, cols) in master_exports.items()]
        )

        write_tables(
            writer,
            "Eksplorasi_SMP",
            [(select_columns(df, cols), title)
            for title, (df, cols) in smp_exports.items()]
        )

    excel_buffer.seek(0)


    # =========================
    # EXPORT PDF (IN MEMORY)
    # =========================
    pdf_buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        leftMargin=1 * inch,
        rightMargin=0.8 * inch,
        topMargin=0.8 * inch,
        bottomMargin=0.8 * inch
    )

    elements = []
    styles = getSampleStyleSheet()


    # =========================
    # STYLE
    # =========================
    cell_left = ParagraphStyle(
        name='cell_left',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_LEFT
    )

    cell_right = ParagraphStyle(
        name='cell_right',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_RIGHT
    )

    header_style = ParagraphStyle(
        name='header_style',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=TA_CENTER
    )

    def build_pdf_table(elements, df, title, doc, styles,
                    header_style, cell_left, cell_right,
                    format_number,
                    landscape_mode=True,
                    add_page_break=True):

        # Optional title
        if title:
            elements.append(Paragraph(title, styles["Title"]))
            elements.append(Spacer(1, 12))

        # =====================
        # Build Data
        # =====================
        data = []

        # Header
        data.append([
            Paragraph(str(col), header_style)
            for col in df.columns
        ])

        # Rows
        for _, row in df.iterrows():
            row_data = [
                Paragraph(str(value), cell_left) if i == 0
                else Paragraph(format_number(value), cell_right)
                for i, value in enumerate(row)
            ]
            data.append(row_data)

        # =====================
        # Column Width
        # =====================
        if landscape_mode:
            page_width, _ = landscape(A4)
        else:
            page_width, _ = A4

        usable_width = page_width - doc.leftMargin - doc.rightMargin
        num_cols = len(df.columns)

        ratio = [2.5] + [1] * (num_cols - 1)
        total_ratio = sum(ratio)

        col_widths = [(r / total_ratio) * usable_width for r in ratio]

        # =====================
        # Create Table
        # =====================
        table = Table(data, colWidths=col_widths, repeatRows=1)

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 24))

        if add_page_break:
            elements.append(PageBreak())

    # =========================
    # BUILD ALL TABLES
    # =========================

    tables_config = [
        (per_kec, "Rekapitulasi Progres SYNC DAPODIK per Kecamatan"),
        (per_kec_NS, "Rekapitulasi Progres SYNC DAPODIK Negeri/Swasta per Kecamatan"),
        (per_bp_NS, "Rekapitulasi Progres SYNC DAPODIK Negeri/Swasta per Jenjang"),
        (per_kec_SMP, "Eksplorasi Progres DAPODIK SMP per Kecamatan"),
        (per_kec_SP_SMP, "Eksplorasi Progres SP dan SYNC SMP per Kecamatan"),
        (per_kec_PDRombel_SMP, "Eksplorasi Progres PD dan Rombel SMP per Kecamatan"),
        (per_kec_GTK_SMP, "Eksplorasi Progres GTK SMP per Kecamatan"),
        (per_kec_ALL_SMP, "Eksplorasi Progres ALL DAPODIK SMP per Kecamatan"),
    ]

    for i, (df, title) in enumerate(tables_config):

        build_pdf_table(
            elements=elements,
            df=df,
            title=title,
            doc=doc,
            styles=styles,
            header_style=header_style,
            cell_left=cell_left,
            cell_right=cell_right,
            format_number=format_number,
            add_page_break=(i < len(tables_config) - 1)  # No page break at last table
        )

    # =========================
    # BUILD PDF
    # =========================
    doc.build(elements)

    # Reset buffer posisi ke awal
    pdf_buffer.seek(0)
    excel_buffer.seek(0)

    # =========================
    # DOWNLOAD SECTION
    # =========================
    st.markdown("### 📥 Download Laporan")

    with st.container():

        col1, col2 = st.columns(2)

        with col1:
            st.download_button(
                label="📊 Download Excel",
                data=excel_buffer,
                file_name="Rekap_Progres_SYNC_DAPODIK.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col2:
            st.download_button(
                label="📄 Download PDF",
                data=pdf_buffer,
                file_name="Rekap_Progres_SYNC_DAPODIK.pdf",
                mime="application/pdf",
                use_container_width=True
            )

