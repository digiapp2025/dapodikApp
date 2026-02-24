import streamlit as st
import pandas as pd
import numpy as np
import io

from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus import PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.markdown(
    """
    <style>
        .block-container {
            padding-top: 2.5rem;
        }
    </style>
    """,
    unsafe_allow_html=True
)

    
# =========================
# FORMAT NUMERIC FUNCTION
# =========================
def format_number(value):
    if pd.isna(value):
        return ""
    if isinstance(value, (int, np.integer)):
        return f"{value:,}"
    if isinstance(value, (float, np.floating)):
        return f"{value:,.2f}"
    return str(value)


# =========================
# AGGRID FUNCTION
# =========================
def show_aggrid(df, key_name):

    gb = GridOptionsBuilder.from_dataframe(df)

    # Pagination
    gb.configure_pagination(
        paginationAutoPageSize=False,
        paginationPageSize=25
    )

    # Default column behavior
    gb.configure_default_column(
        filter=True,
        sortable=True,
        resizable=True,
        floatingFilter=True
    )

    # Sidebar (filter & column selector)
    gb.configure_side_bar()
    
    numeric_cols = df.select_dtypes(include=["int64", "float64"]).columns

    for col in numeric_cols:
        gb.configure_column(
            col,
            type=["numericColumn"],
            valueFormatter="x.toLocaleString()"
        )

    # Freeze kolom Kecamatan jika ada
    if "Kecamatan" in df.columns:
        gb.configure_column(
            "Kecamatan",
            pinned="left"
        )
    
    # Freeze kolom pertama (index 0)
    if len(df.columns) > 0:
        first_col = df.columns[0]

        gb.configure_column(
            first_col,
            pinned="left"
        )

    # Grid options
    gb.configure_grid_options(
        domLayout="normal",
        enableRangeSelection=True
    )

    grid_options = gb.build()

    AgGrid(
        df,
        gridOptions=grid_options,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=True,
        theme="streamlit",
        key=key_name,
        height=500
    )



# =========================
# STREAMLIT UI
# =========================
st.title("Rekap Progres SYNC DAPODIK")

uploaded_file = st.file_uploader("Upload File Excel", type=["xlsx"])

if uploaded_file:

    # =========================
    # BACA FILE
    # =========================
    dapodik = pd.read_excel(uploaded_file, sheet_name="Master")

    dapodik["BP"] = dapodik["BP"].astype(str).str.strip().str.upper()
    dapodik["Status"] = dapodik["Status"].astype(str).str.strip().str.upper()
    dapodik["Last Sync"] = dapodik["Last Sync"].astype(str).str.strip().str.lower()

    dapodik_filtered = dapodik.loc[~dapodik["BP"].isin(["SMA", "SMK", "SLB"])].copy()
    dapodik_filtered_SMP = dapodik.loc[dapodik["BP"].isin(["SMP"])].copy()

    dapodik_filtered["Status_SYNC"] = dapodik_filtered["Last Sync"].apply(
        lambda x: "Belum" if x == "belum kirim" else "Sudah"
    )
    dapodik_filtered_SMP["Status_SYNC"] = dapodik_filtered_SMP["Last Sync"].apply(
        lambda x: "Belum" if x == "belum kirim" else "Sudah"
    )

    dapodik_filtered["Jml_GTK"] = (
        dapodik_filtered["Guru"] + dapodik_filtered["Tendik"]
    )
    dapodik_filtered_SMP["Jml_GTK"] = (
        dapodik_filtered_SMP["Guru"] + dapodik_filtered_SMP["Tendik"]
    )

    # =========================
    # KOLOM NEGERI / SWASTA
    # =========================
    for col in ["NPSN", "PD", "Rombel", "Guru", "Tendik", "Jml_GTK"]:
        if col == "NPSN":
            dapodik_filtered[f"{col}_Negeri"] = np.where(
                dapodik_filtered["Status"] == "NEGERI",
                dapodik_filtered[col],
                None
            )
            dapodik_filtered[f"{col}_Swasta"] = np.where(
                dapodik_filtered["Status"] == "SWASTA",
                dapodik_filtered[col],
                None
            )
        else:
            dapodik_filtered[f"{col}_Negeri"] = np.where(
                dapodik_filtered["Status"] == "NEGERI",
                dapodik_filtered[col],
                0
            )
            dapodik_filtered[f"{col}_Swasta"] = np.where(
                dapodik_filtered["Status"] == "SWASTA",
                dapodik_filtered[col],
                0
            )
    
    # =========================
    # KOLOM NEGERI / SWASTA JENJANG SMP
    # =========================
    for col_SMP in ["NPSN", "PD", "Rombel", "Guru", "Tendik", "Jml_GTK"]:
        if col_SMP == "NPSN":
            dapodik_filtered_SMP[f"{col_SMP}_Negeri"] = np.where(
                dapodik_filtered_SMP["Status"] == "NEGERI",
                dapodik_filtered_SMP[col_SMP],
                None
            )
            dapodik_filtered_SMP[f"{col_SMP}_Swasta"] = np.where(
                dapodik_filtered_SMP["Status"] == "SWASTA",
                dapodik_filtered_SMP[col_SMP],
                None
            )
        else:
            dapodik_filtered_SMP[f"{col_SMP}_Negeri"] = np.where(
                dapodik_filtered_SMP["Status"] == "NEGERI",
                dapodik_filtered_SMP[col_SMP],
                0
            )
            dapodik_filtered_SMP[f"{col_SMP}_Swasta"] = np.where(
                dapodik_filtered_SMP["Status"] == "SWASTA",
                dapodik_filtered_SMP[col_SMP],
                0
            )


    # =========================
    # AGREGASI
    # =========================
    per_kec = (
        dapodik_filtered
        .groupby("Kecamatan")
        .agg(
            SP=("NPSN", "nunique"),
            Sudah_SYNC=("Status_SYNC", lambda x: (x == "Sudah").sum()),
            Belum_SYNC=("Status_SYNC", lambda x: (x == "Belum").sum()),
            PD=("PD", "sum"),
            Rombel=("Rombel", "sum"),
            Guru=("Guru", "sum"),
            Tendik=("Tendik", "sum"),
            Jml_GTK=("Jml_GTK", "sum")
        )
        .reset_index()
        .sort_values("Kecamatan")
    )

    per_kec_NS = (
        dapodik_filtered
        .groupby("Kecamatan")
        .agg(
            SP_Negeri=("NPSN_Negeri", "nunique"),
            SP_Swasta=("NPSN_Swasta", "nunique"),
            Jml_SP=("NPSN", "nunique"),
            
            PD_Negeri=("PD_Negeri", "sum"),
            PD_Swasta=("PD_Swasta", "sum"),
            Jml_PD=("PD", "sum"),
            
            Rombel_Negeri=("Rombel_Negeri", "sum"),
            Rombel_Swasta=("Rombel_Swasta", "sum"),
            Jml_Rombel=("Rombel", "sum"),
            
            Guru_Negeri=("Guru_Negeri", "sum"),
            Guru_Swasta=("Guru_Swasta", "sum"),
            Jml_Guru=("Guru", "sum"),
            
            Tendik_Negeri=("Tendik_Negeri", "sum"),
            Tendik_Swasta=("Tendik_Swasta", "sum"),
            Jml_Tendik=("Tendik", "sum"),
            
            Jml_GTK_Negeri=("Jml_GTK_Negeri", "sum"),
            Jml_GTK_Swasta=("Jml_GTK_Swasta", "sum"),
            Jml_GTK=("Jml_GTK", "sum")
        )
        .reset_index()
        .sort_values("Kecamatan")
    )

    per_bp_NS = (
        dapodik_filtered
        .groupby("BP")
        .agg(
            SP_Negeri=("NPSN_Negeri", "nunique"),
            SP_Swasta=("NPSN_Swasta", "nunique"),
            Jml_SP=("NPSN", "nunique"),
            
            PD_Negeri=("PD_Negeri", "sum"),
            PD_Swasta=("PD_Swasta", "sum"),
            Jml_PD=("PD", "sum"),
            
            Rombel_Negeri=("Rombel_Negeri", "sum"),
            Rombel_Swasta=("Rombel_Swasta", "sum"),
            Jml_Rombel=("Rombel", "sum"),
            
            Guru_Negeri=("Guru_Negeri", "sum"),
            Guru_Swasta=("Guru_Swasta", "sum"),
            Jml_Guru=("Guru", "sum"),
            
            Tendik_Negeri=("Tendik_Negeri", "sum"),
            Tendik_Swasta=("Tendik_Swasta", "sum"),
            Jml_Tendik=("Tendik", "sum"),
            
            Jml_GTK_Negeri=("Jml_GTK_Negeri", "sum"),
            Jml_GTK_Swasta=("Jml_GTK_Swasta", "sum"),
            Jml_GTK=("Jml_GTK", "sum")
        )
        .reset_index()
        .sort_values("BP")
    )
    
    # =========================
    # AGREGASI JENJANG SMP
    # =========================
    
    per_kec_SMP = (
        dapodik_filtered_SMP
        .groupby("Kecamatan")
        .agg(
            SP=("NPSN", "nunique"),
            Sudah_SYNC=("Status_SYNC", lambda x: (x == "Sudah").sum()),
            Belum_SYNC=("Status_SYNC", lambda x: (x == "Belum").sum()),
            PD=("PD", "sum"),
            Rombel=("Rombel", "sum"),
            Guru=("Guru", "sum"),
            Tendik=("Tendik", "sum"),
            Jml_GTK=("Jml_GTK", "sum")
        )
        .reset_index()
        .sort_values("Kecamatan")
    )
    
    per_kec_SP_SMP = (
        dapodik_filtered_SMP
        .groupby("Kecamatan")
        .agg(
            Jml_SP=("NPSN", "nunique"),
            SP_Negeri=("NPSN_Negeri", "nunique"),
            SP_Swasta=("NPSN_Swasta", "nunique"),            
            
            Sudah_SYNC=("Status_SYNC", lambda x: (x == "Sudah").sum()),
            Belum_SYNC=("Status_SYNC", lambda x: (x == "Belum").sum())
        )
        .reset_index()
        .sort_values("Kecamatan")
    )
    
    per_kec_PDRombel_SMP = (
        dapodik_filtered_SMP
        .groupby("Kecamatan")
        .agg(
            PD_Negeri=("PD_Negeri", "sum"),
            PD_Swasta=("PD_Swasta", "sum"),
            Jml_PD=("PD", "sum"),
            
            Rombel_Negeri=("Rombel_Negeri", "sum"),
            Rombel_Swasta=("Rombel_Swasta", "sum"),
            Jml_Rombel=("Rombel", "sum")
        )
        .reset_index()
        .sort_values("Kecamatan")
    )
    
    per_kec_GTK_SMP = (
        dapodik_filtered_SMP
        .groupby("Kecamatan")
        .agg(
            Jml_GTK=("Jml_GTK", "sum"),
            
            Guru_Negeri=("Guru_Negeri", "sum"),
            Guru_Swasta=("Guru_Swasta", "sum"),
            Jml_Guru=("Guru", "sum"),
            
            Tendik_Negeri=("Tendik_Negeri", "sum"),
            Tendik_Swasta=("Tendik_Swasta", "sum"),
            Jml_Tendik=("Tendik", "sum")            
        )
        .reset_index()
        .sort_values("Kecamatan")
    )
    
    per_kec_ALL_SMP = (
        dapodik_filtered_SMP
        .groupby("Kecamatan")
        .agg(            
            SP_Negeri=("NPSN_Negeri", "nunique"),
            SP_Swasta=("NPSN_Swasta", "nunique"),
            
            PD_Negeri=("PD_Negeri", "sum"),
            PD_Swasta=("PD_Swasta", "sum"),            
            
            Rombel_Negeri=("Rombel_Negeri", "sum"),
            Rombel_Swasta=("Rombel_Swasta", "sum"),                      
            
            Guru_Negeri=("Guru_Negeri", "sum"),
            Guru_Swasta=("Guru_Swasta", "sum"),            
            
            Tendik_Negeri=("Tendik_Negeri", "sum"),
            Tendik_Swasta=("Tendik_Swasta", "sum"),
            
            Jml_SP=("NPSN", "nunique"),
            Jml_PD=("PD", "sum"),
            Jml_Rombel=("Rombel", "sum"),
            Jml_Guru=("Guru", "sum"),
            Jml_Tendik=("Tendik", "sum"),
            Jml_GTK=("Jml_GTK", "sum")            
        )
        .reset_index()
        .sort_values("Kecamatan")
    )

    # =========================
    # TOTAL ROW
    # =========================
    def add_total_row(df, label_column):
        df = df.copy()

        # 🔹 Hapus kolom auto id jika ada
        df = df.loc[:, ~df.columns.str.contains("::auto_unique_id::", case=False)]

        # 🔹 Ambil hanya kolom numerik untuk dijumlahkan
        numeric_cols = df.select_dtypes(include="number").columns

        # 🔹 Hitung total
        total_values = df[numeric_cols].sum()

        # 🔹 Buat row total
        total_row = {col: "" for col in df.columns}
        total_row[label_column] = "TOTAL"

        for col in numeric_cols:
            total_row[col] = total_values[col]

        # 🔹 Append ke dataframe
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        return df

    per_kec_final = add_total_row(per_kec, "Kecamatan")
    per_kec_NS_final = add_total_row(per_kec_NS, "Kecamatan")
    per_bp_NS_final = add_total_row(per_bp_NS, "BP")
    
    per_kec_SMP_final = add_total_row(per_kec_SMP, "Kecamatan")
    per_kec_SP_SMP_final = add_total_row(per_kec_SP_SMP, "Kecamatan")
    per_kec_PDRombel_SMP_final = add_total_row(per_kec_PDRombel_SMP, "Kecamatan")
    per_kec_GTK_SMP_final = add_total_row(per_kec_GTK_SMP, "Kecamatan")
    per_kec_ALL_SMP_final = add_total_row(per_kec_ALL_SMP, "Kecamatan")


    st.success("Data berhasil diproses!")
    
    # st.write("Kolom sebelum export:", per_kec_final.columns.tolist()) DEBUG

    
    # =========================
    # KPI DASHBOARD
    # =========================

    # total_sp = per_kec["SP"].sum()
    # total_sudah = per_kec["Sudah_SYNC"].sum()
    # total_belum = per_kec["Belum_SYNC"].sum()
    # persen_sync = (total_sudah / total_sp * 100) if total_sp else 0
    
    total_sp = dapodik_filtered["NPSN"].nunique()
    total_sudah = dapodik_filtered.loc[
        dapodik_filtered["Status_SYNC"] == "Sudah", "NPSN"
    ].nunique()

    total_belum = total_sp - total_sudah

    persen_sync = (total_sudah / total_sp * 100) if total_sp > 0 else 0

    col1, col2, col3, col4 = st.columns(4)

    col1.metric("Total Satuan Pendidikan", f"{total_sp:,}")
    col2.metric("Sudah SYNC", f"{total_sudah:,}")
    col3.metric("Belum SYNC", f"{total_belum:,}")
    col4.metric("Persentase SYNC", f"{persen_sync:.2f}%")
    
    # =========================
    # GRAFIK PROGRES SYNC
    # =========================

    st.subheader("📊 Grafik Progres SYNC per Kecamatan")

    chart_data = per_kec[["Kecamatan", "Sudah_SYNC", "Belum_SYNC"]]
    chart_data = chart_data.set_index("Kecamatan")

    st.bar_chart(chart_data)
    
    # =========================
    # RANKING KECAMATAN
    # =========================

    ranking = per_kec.copy()
    ranking["Persentase_SYNC"] = (
        ranking["Sudah_SYNC"] / ranking["SP"] * 100
    ).round(2)

    ranking = ranking.sort_values(
        "Persentase_SYNC", ascending=False
    )

    st.subheader("🏆 Ranking Kecamatan Berdasarkan Persentase SYNC")

    st.dataframe(
        ranking[["Kecamatan", "SP", "Sudah_SYNC", "Belum_SYNC", "Persentase_SYNC"]],
        use_container_width=True
    )
    
    # =========================
    # TABS UNTUK TABEL
    # =========================

    tab1, tab2, tab3 = st.tabs([
        "📊 Rekap Per Kecamatan",
        "🏫 Rekap Negeri/Swasta Per Kecamatan",
        "📌 Rekap Negeri/Swasta Per BP"
    ])

    with tab1:
        show_aggrid(per_kec_final, "grid_kec")

    with tab2:
        show_aggrid(per_kec_NS_final, "grid_kec_ns")

    with tab3:
        show_aggrid(per_bp_NS_final, "grid_bp_ns")


    # =========================
    # TABS UNTUK TABEL JENJANG SMP
    # =========================

    tab_smp_1, tab_smp_2, tab_smp_3, tab_smp_4, tab_smp_5 = st.tabs([
        "📊 DAPODIK SMP",
        "🏫 SP dan SYNC SMP",
        "📌 PD dan Rombel SMP",
        "📌 Progres GTK SMP",
        "📌 ALL DAPODIK SMP"
    ])

    with tab_smp_1:
        show_aggrid(per_kec_SMP_final, "grid_kec_smp")

    with tab_smp_2:
        show_aggrid(per_kec_SP_SMP_final, "grid_kec_sp_smp")

    with tab_smp_3:
        show_aggrid(per_kec_PDRombel_SMP_final, "grid_kec_pdrombel_smp")
        
    with tab_smp_4:
        show_aggrid(per_kec_GTK_SMP_final, "grid_kec_gtk_smp")
    
    with tab_smp_5:
        show_aggrid(per_kec_ALL_SMP_final, "grid_kec_all_smp")
        
            


    # =========================================
    # GLOBAL STYLE CONFIG
    # =========================================
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(size=14, bold=True)
    TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    NUMBER_FORMAT_INTEGER = '#,##0'
    NUMBER_FORMAT_DECIMAL = '#,##0.00'
    NUMBER_FORMAT_PERCENT = '0.00%'


    # =========================================
    # STYLE TABLE FUNCTION
    # =========================================
    def style_table(ws, start_row, df, title=None):
        """
        start_row : baris pertama tabel (1-based Excel)
        df        : dataframe
        """

        n_cols = len(df.columns)
        header_row = start_row
        data_start = header_row + 1
        data_end = data_start + len(df) - 1
        total_row = data_end

        # =========================
        # TITLE
        # =========================
        if title:
            ws.insert_rows(start_row)
            ws.cell(row=start_row, column=1, value=title)
            ws.cell(row=start_row, column=1).font = TITLE_FONT
            ws.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=start_row,
                end_column=n_cols
            )
            header_row += 1
            data_start += 1
            data_end += 1
            total_row += 1

        # =========================
        # HEADER STYLE
        # =========================
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # =========================
        # DATA + TOTAL STYLE
        # =========================
        for row in ws.iter_rows(
            min_row=data_start,
            max_row=data_end,
            min_col=1,
            max_col=n_cols
            
        ):
            for i, cell in enumerate(row):
                col_name = df.columns[i]
                
                cell.border = THIN_BORDER                
                
                if i == 0:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="right")
                
                # =========================
                # FORMAT NUMERIC OTOMATIS
                # =========================
                if pd.api.types.is_integer_dtype(df[col_name]):
                    cell.number_format = NUMBER_FORMAT_INTEGER

                elif pd.api.types.is_float_dtype(df[col_name]):

                    # Jika kolom persentase
                    if "%" in col_name.lower() or "persen" in col_name.lower():
                        cell.number_format = NUMBER_FORMAT_PERCENT
                    else:
                        cell.number_format = NUMBER_FORMAT_DECIMAL

        # Highlight TOTAL row
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)


        # =========================
        # AUTO WIDTH (Only Table Range)
        # =========================
        for col in range(1, n_cols + 1):
            column_letter = get_column_letter(col)
            max_len = 0

            for row in range(header_row, data_end + 1):
                value = ws.cell(row=row, column=col).value
                if value:
                    max_len = max(max_len, len(str(value)))

            ws.column_dimensions[column_letter].width = max_len + 3


    # =========================================
    # WRITE MULTIPLE TABLES
    # =========================================
    def write_tables(writer, sheet_name, tables, spacing=4):

        ws = writer.book.create_sheet(sheet_name)
        writer.sheets[sheet_name] = ws

        current_row = 1

        for df, title in tables:

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=current_row - 1
            )

            style_table(
                ws=ws,
                start_row=current_row,
                df=df,
                title=title
            )

            # Update posisi berikutnya
            current_row += len(df) + spacing + 2


    # =========================================
    # EXPORT EXECUTION
    # =========================================
    excel_buffer = io.BytesIO()
    
    export_kec = per_kec_final[
        ["Kecamatan","SP","Sudah_SYNC","Belum_SYNC","PD","Rombel","Guru","Tendik","Jml_GTK"]
    ]
    
    export_kec_NS = per_kec_NS_final[
        ["Kecamatan", "SP_Negeri", "SP_Swasta", "Jml_SP", "PD_Negeri", "PD_Swasta", "Jml_PD", "Rombel_Negeri", "Rombel_Swasta", "Jml_Rombel", 
        "Guru_Negeri", "Guru_Swasta", "Jml_Guru", "Tendik_Negeri", "Tendik_Swasta", "Jml_Tendik", "Jml_GTK_Negeri", "Jml_GTK_Swasta", "Jml_GTK"]
    ]
    
    export_bp_NS = per_bp_NS_final[
        ["BP", "SP_Negeri", "SP_Swasta", "Jml_SP", "PD_Negeri", "PD_Swasta", "Jml_PD", "Rombel_Negeri", "Rombel_Swasta", "Jml_Rombel", 
        "Guru_Negeri", "Guru_Swasta", "Jml_Guru", "Tendik_Negeri", "Tendik_Swasta", "Jml_Tendik", "Jml_GTK_Negeri", "Jml_GTK_Swasta", "Jml_GTK"]
    ]
    
    # =============
    # EKSPLORASI
    # =============
    
    export_kec_SMP = per_kec_SMP_final[
        ["Kecamatan","SP","Sudah_SYNC","Belum_SYNC","PD","Rombel","Guru","Tendik","Jml_GTK"]
    ]
        
    export_kec_SP_SMP = per_kec_SP_SMP_final[
        ["Kecamatan", "Jml_SP", "SP_Negeri", "SP_Swasta", "Sudah_SYNC","Belum_SYNC"]
    ]
        
    export_kec_PDRombel_SMP = per_kec_PDRombel_SMP_final[
        ["Kecamatan", "PD_Negeri", "PD_Swasta", "Jml_PD", "Rombel_Negeri", "Rombel_Swasta", "Jml_Rombel"]
    ]
        
    export_kec_GTK_SMP = per_kec_GTK_SMP_final[
        ["Kecamatan", "Jml_GTK", "Guru_Negeri", "Guru_Swasta", "Jml_Guru", "Tendik_Negeri", "Tendik_Swasta", "Jml_Tendik"]
    ]
        
    export_kec_ALL_SMP = per_kec_ALL_SMP_final[
        ["Kecamatan", "SP_Negeri", "SP_Swasta", "PD_Negeri", "PD_Swasta", "Rombel_Negeri", "Rombel_Swasta", "Guru_Negeri", "Guru_Swasta", "Jml_Guru", 
        "Tendik_Negeri", "Tendik_Swasta", "Jml_SP", "Jml_PD", "Jml_Rombel", "Jml_Tendik", "Jml_GTK"]
    ]
    
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:

        write_tables(
            writer,
            "Rekap_Master",
            [
                (export_kec, "Rekapitulasi Progres SYNC DAPODIK per Kecamatan"),
                (export_kec_NS, "Rekapitulasi SYNC Negeri/Swasta per Kecamatan"),
                (export_bp_NS, "Rekapitulasi SYNC Negeri/Swasta per Jenjang"),
            ]
        )

        write_tables(
            writer,
            "Eksplorasi_SMP",
            [
                (export_kec_SMP, "Eksplorasi Progres SMP per Kecamatan"),
                (export_kec_SP_SMP, "Eksplorasi SP & SYNC SMP"),
                (export_kec_PDRombel_SMP, "Eksplorasi PD & Rombel SMP"),
                (export_kec_GTK_SMP, "Eksplorasi GTK SMP"),
                (export_kec_ALL_SMP, "Eksplorasi ALL SMP"),
            ]
        )

    excel_buffer.seek(0)



    # =========================
    # EXPORT PDF (IN MEMORY)
    # =========================
    pdf_buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        leftMargin=1 * inch,
        rightMargin=0.8 * inch,
        topMargin=0.8 * inch,
        bottomMargin=0.8 * inch
    )

    elements = []
    styles = getSampleStyleSheet()


    # =========================
    # STYLE
    # =========================
    cell_left = ParagraphStyle(
        name='cell_left',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_LEFT
    )

    cell_right = ParagraphStyle(
        name='cell_right',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_RIGHT
    )

    header_style = ParagraphStyle(
        name='header_style',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=TA_CENTER
    )

    def build_pdf_table(elements, df, title, doc, styles,
                    header_style, cell_left, cell_right,
                    format_number,
                    landscape_mode=True,
                    add_page_break=True):

        # Optional title
        if title:
            elements.append(Paragraph(title, styles["Title"]))
            elements.append(Spacer(1, 12))

        # =====================
        # Build Data
        # =====================
        data = []

        # Header
        data.append([
            Paragraph(str(col), header_style)
            for col in df.columns
        ])

        # Rows
        for _, row in df.iterrows():
            row_data = [
                Paragraph(str(value), cell_left) if i == 0
                else Paragraph(format_number(value), cell_right)
                for i, value in enumerate(row)
            ]
            data.append(row_data)

        # =====================
        # Column Width
        # =====================
        if landscape_mode:
            page_width, _ = landscape(A4)
        else:
            page_width, _ = A4

        usable_width = page_width - doc.leftMargin - doc.rightMargin
        num_cols = len(df.columns)

        ratio = [2.5] + [1] * (num_cols - 1)
        total_ratio = sum(ratio)

        col_widths = [(r / total_ratio) * usable_width for r in ratio]

        # =====================
        # Create Table
        # =====================
        table = Table(data, colWidths=col_widths, repeatRows=1)

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 24))

        if add_page_break:
            elements.append(PageBreak())

    # =========================
    # BUILD ALL TABLES
    # =========================

    tables_config = [
        (export_kec, "Rekapitulasi Progres SYNC DAPODIK per Kecamatan"),
        (export_kec_NS, "Rekapitulasi Progres SYNC DAPODIK Negeri/Swasta per Kecamatan"),
        (export_bp_NS, "Rekapitulasi Progres SYNC DAPODIK Negeri/Swasta per Jenjang"),
        (export_kec_SMP, "Eksplorasi Progres DAPODIK SMP per Kecamatan"),
        (export_kec_SP_SMP, "Eksplorasi Progres SP dan SYNC SMP per Kecamatan"),
        (export_kec_PDRombel_SMP, "Eksplorasi Progres PD dan Rombel SMP per Kecamatan"),
        (export_kec_GTK_SMP, "Eksplorasi Progres GTK SMP per Kecamatan"),
        (export_kec_ALL_SMP, "Eksplorasi Progres ALL DAPODIK SMP per Kecamatan"),
    ]

    for i, (df, title) in enumerate(tables_config):

        build_pdf_table(
            elements=elements,
            df=df,
            title=title,
            doc=doc,
            styles=styles,
            header_style=header_style,
            cell_left=cell_left,
            cell_right=cell_right,
            format_number=format_number,
            add_page_break=(i < len(tables_config) - 1)  # No page break at last table
        )

    # =========================
    # BUILD PDF
    # =========================
    doc.build(elements)

    # Reset buffer posisi ke awal
    pdf_buffer.seek(0)
    excel_buffer.seek(0)

    # =========================
    # DOWNLOAD SECTION
    # =========================
    st.markdown("### 📥 Download Laporan")

    with st.container():

        col1, col2 = st.columns(2)

        with col1:
            st.download_button(
                label="📊 Download Excel",
                data=excel_buffer,
                file_name="Rekap_Progres_SYNC_DAPODIK.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col2:
            st.download_button(
                label="📄 Download PDF",
                data=pdf_buffer,
                file_name="Rekap_Progres_SYNC_DAPODIK.pdf",
                mime="application/pdf",
                use_container_width=True
            )

